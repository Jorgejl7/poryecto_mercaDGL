import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog  # Importar simpledialog
from tkinter import ttk
from PIL import Image, ImageTk
import cv2
import pandas as pd
import datetime
from pyzbar.pyzbar import decode
from openpyxl import load_workbook # noqa: F401
from openpyxl.styles import Alignment

def cargar_imagen(entry_codigo, lbl_imagen_destino, entry_nombre):
    file_path = filedialog.askopenfilename(filetypes=[("Imágenes", "*.png;*.jpg;*.jpeg")])
    if not file_path:
        return
    mostrar_imagen(file_path, entry_codigo, lbl_imagen_destino, entry_nombre)

def capturar_imagen(entry_codigo, lbl_imagen_destino, entry_nombre):
    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        messagebox.showerror("Error", "No se pudo abrir la cámara.")
        return

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        barcode = decode(frame)
        for b in barcode:
            codigo_barras = b.data.decode("utf-8")
            cap.release()
            cv2.destroyAllWindows()
            entry_codigo.delete(0, tk.END)
            entry_codigo.insert(0, codigo_barras)
            file_path = "captura.jpg"
            cv2.imwrite(file_path, frame)
            mostrar_imagen(file_path, entry_codigo, lbl_imagen_destino, entry_nombre)
            return
        
        cv2.imshow("Escanea el código de barras", frame)
        if cv2.waitKey(1) & 0xFF == 27:
            break

    cap.release()
    cv2.destroyAllWindows()

def mostrar_imagen(file_path, entry_codigo, lbl_imagen_destino, entry_nombre):
    img = Image.open(file_path)
    img = img.resize((250, 150))
    img_tk = ImageTk.PhotoImage(img)
    lbl_imagen_destino.config(image=img_tk)
    lbl_imagen_destino.image = img_tk
    
    img_cv = cv2.imread(file_path)
    img_cv = cv2.resize(img_cv, (400, 300))
    barcode = decode(img_cv)
    if barcode:
        codigo = barcode[0].data.decode("utf-8")
        entry_codigo.delete(0, tk.END)
        entry_codigo.insert(0, codigo)
        # Mover automáticamente el cursor al campo Nombre
        entry_nombre.focus()

def guardar_datos(lista, entrys, lbl_imagen):
    try:
        nombre = entrys[0].get()
        cantidad = int(entrys[1].get())
        precio = float(entrys[2].get())
        codigo = entrys[3].get()
    except ValueError:
        messagebox.showwarning("Advertencia", "Cantidad debe ser entero y precio un número válido.")
        return

    fecha = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if not (nombre and cantidad and precio and codigo):
        messagebox.showwarning("Advertencia", "Completa todos los campos.")
        return

    lista.append([nombre, cantidad, f"Q{precio}", fecha, codigo])
    messagebox.showinfo("Éxito", "Registro guardado correctamente.")
    for e in entrys:
        e.delete(0, tk.END)
    lbl_imagen.config(image="")
    
    # Enfocar automáticamente en el campo de Código de Barras
    entrys[3].focus()

def exportar_excel(lista):
    if not lista:
        messagebox.showwarning("Advertencia", "No hay datos para exportar.")
        return
    
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Archivos Excel", "*.xlsx")])
    if not file_path:
        return

    # Crear DataFrame y exportar a Excel
    df = pd.DataFrame(lista, columns=["Nombre", "Cantidad", "Precio", "Fecha", "Código de Barras"])
    
    # Guardar el DataFrame en Excel
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Registros")

        # Acceder a la hoja y aplicar estilos
        workbook = writer.book
        sheet = workbook["Registros"]
        
        # Centrar todo el contenido
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Ajustar el tamaño de las columnas
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Obtener la letra de la columna
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

    messagebox.showinfo("Éxito", f"Datos exportados a {file_path}")

def on_codigo_ingresado(event, entry_codigo, entry_nombre):
    codigo = entry_codigo.get().strip()
    if codigo:
        entry_nombre.focus()

def leer_codigo_barra(entry_codigo, entry_nombre):
    entry_codigo.focus()
    entry_codigo.bind("<Return>", lambda event: on_codigo_ingresado(event, entry_codigo, entry_nombre))
    
def crear_interfaz_registro(titulo, lista, x_offset, y_offset):
    ventana = tk.Toplevel(root)
    ventana.title(titulo)
    ventana.geometry(f"400x500+{x_offset}+{y_offset}")
    ventana.configure(bg="#e6f7ff")
    
    frame = tk.Frame(ventana, bg="#e6f7ff")
    frame.pack(expand=True)
    
    entrys = []
    etiquetas = ["Nombre Producto:", "Cantidad:", "Precio (Q):", "Código de Barras:"]
    for i, text in enumerate(etiquetas):
        tk.Label(frame, text=text, bg="#e6f7ff").grid(row=i, column=0, padx=10, pady=5)
        entry = tk.Entry(frame)
        entry.grid(row=i, column=1, padx=10, pady=5)
        entrys.append(entry)
    
    leer_codigo_barra(entrys[3], entrys[0])
    
    lbl_imagen = tk.Label(frame, bg="#e6f7ff")
    lbl_imagen.grid(row=5, column=0, columnspan=2, pady=10)
    
    tk.Button(frame, text="Seleccionar Imagen", command=lambda: cargar_imagen(entrys[3], lbl_imagen, entrys[0])).grid(row=6, column=0, padx=5, pady=10)
    tk.Button(frame, text="Capturar desde Cámara", command=lambda: capturar_imagen(entrys[3], lbl_imagen, entrys[0])).grid(row=6, column=1, padx=5, pady=10)
    tk.Button(frame, text="Guardar", command=lambda: guardar_datos(lista, entrys, lbl_imagen)).grid(row=7, column=0, padx=5, pady=10)
    tk.Button(frame, text="Exportar a Excel", command=lambda: exportar_excel(lista)).grid(row=7, column=1, padx=5, pady=10)
    
    # Agregar el botón "Historial"
    tk.Button(frame, text="Historial", command=lambda: mostrar_historial(lista, "Productos" if titulo == "Registro de Productos MercaDGL" else "Ventas")).grid(row=8, column=0, columnspan=2, pady=10)
    
    tk.Button(frame, text="Cerrar", command=ventana.destroy).grid(row=9, column=0, columnspan=2, pady=10)

def mostrar_historial(lista, tipo="Productos"):
    # Crear una ventana para mostrar el historial
    historial_ventana = tk.Toplevel(root)
    historial_ventana.title(f"Historial de {tipo}")
    historial_ventana.geometry("800x500")
    historial_ventana.configure(bg="#f0f8ff")
    
    # Crear un frame para contener la tabla y los scrollbars
    frame_historial = tk.Frame(historial_ventana, bg="#f0f8ff")
    frame_historial.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
    
    # Crear un Treeview para mostrar la tabla
    treeview = ttk.Treeview(frame_historial, columns=("Nombre", "Cantidad", "Precio", "Fecha", "Código"), show="headings", height=15)
    treeview.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    
    # Crear scrollbar vertical
    scrollbar = ttk.Scrollbar(frame_historial, orient="vertical", command=treeview.yview)
    scrollbar.pack(side=tk.RIGHT, fill="y")
    treeview.configure(yscrollcommand=scrollbar.set)
    
    # Definir las columnas
    treeview.heading("Nombre", text="Nombre Producto" if tipo == "Productos" else "Nombre Venta")
    treeview.heading("Cantidad", text="Cantidad")
    treeview.heading("Precio", text="Precio (Q)")
    treeview.heading("Fecha", text="Fecha")
    treeview.heading("Código", text="Código de Barras")
    
    # Definir el tamaño de las columnas
    treeview.column("Nombre", width=220, anchor="center")
    treeview.column("Cantidad", width=100, anchor="center")
    treeview.column("Precio", width=120, anchor="center")
    treeview.column("Fecha", width=140, anchor="center")
    treeview.column("Código", width=120, anchor="center")
    
    # Agregar los datos al Treeview
    for registro in lista:
        treeview.insert("", tk.END, values=(registro[0], registro[1], registro[2], registro[3], registro[4]))

    # Estilo para filas alternas (par e impar)
    def style_treeview():
        for i, row in enumerate(treeview.get_children()):
            if i % 2 == 0:
                treeview.item(row, tags='even')
            else:
                treeview.item(row, tags='odd')
        
        treeview.tag_configure('even', background="#e6f7ff")
        treeview.tag_configure('odd', background="#ffffff")

    style_treeview()

    def modificar_registro(treeview, lista):
        selected_item = treeview.selection()
        if not selected_item:
            messagebox.showwarning("Advertencia", "Selecciona un registro para modificar.")
            return
    
        item = treeview.item(selected_item)
        values = item['values']
    
        # Solicitar nuevo nombre, cantidad y precio
        nuevo_nombre = simpledialog.askstring("Modificar Nombre", "Nuevo nombre del producto:", initialvalue=values[0])
        if not nuevo_nombre:
            return
    
        nueva_cantidad = simpledialog.askinteger("Modificar Cantidad", "Nueva cantidad:", initialvalue=int(values[1]))
        if nueva_cantidad is None:
            return
    
        nuevo_precio = simpledialog.askfloat("Modificar Precio", "Nuevo precio:", initialvalue=float(values[2][1:]))  
        if nuevo_precio is None:
            return
    
        nueva_fecha = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
        # Actualizar la lista en memoria
        for i, registro in enumerate(lista):
            if registro[3] == values[3]:  # Comparar por código de barras para asegurar coincidencia
                lista[i] = [nuevo_nombre, nueva_cantidad, f"Q{nuevo_precio}", values[3], nueva_fecha]
                break
    
        # Actualizar el Treeview
        treeview.item(selected_item, values=(nuevo_nombre, nueva_cantidad, f"Q{nuevo_precio}", values[3], nueva_fecha))
        messagebox.showinfo("Éxito", "Registro modificado correctamente.")
    
    def eliminar_registro(treeview, lista):
        selected_item = treeview.selection()
        if not selected_item:
            messagebox.showwarning("Advertencia", "Selecciona un registro para eliminar.")
            return
    
        confirmacion = messagebox.askyesno("Confirmar eliminación", "¿Estás seguro de que deseas eliminar este registro?")
        if not confirmacion:
            return
    
        item = treeview.item(selected_item)
        values = item['values']
    
        # Eliminar de la lista en memoria
        lista[:] = [registro for registro in lista if registro[3] != values[3]]  # Comparar por código de barras
    
        # Eliminar del Treeview
        treeview.delete(selected_item)
        messagebox.showinfo("Éxito", "Registro eliminado correctamente.")


    def cerrar_historial():
        # Cambiado para que solo cierre la ventana sin eliminar o modificar nada
        historial_ventana.destroy()
    
    
    # Agregar botones de acción
    tk.Button(historial_ventana, text="Modificar", command=lambda:modificar_registro(treeview,lista)).pack(pady=10)
    tk.Button(historial_ventana, text="Eliminar", command=lambda:eliminar_registro(treeview,lista)).pack(pady=10)
    tk.Button(historial_ventana, text="Cerrar Historial", command=cerrar_historial).pack(pady=10)

def abrir_registro_productos():
    crear_interfaz_registro("Registro de Productos MercaDGL", productos, 420, 100)

def abrir_registro_ventas():
    crear_interfaz_registro("Registro de Ventas MercaDGL", ventas, 840, 100)

# Inicializamos la ventana principal
root = tk.Tk()
root.title("MercaDGL")
root.geometry("400x500")
root.configure(bg="#e6f7ff")

productos = []
ventas = []

frame_main = tk.Frame(root, bg="#e6f7ff")
frame_main.pack(expand=True)

tk.Button(frame_main, text="Registro de Productos", command=abrir_registro_productos).pack(pady=10)
tk.Button(frame_main, text="Registro de Ventas", command=abrir_registro_ventas).pack(pady=10)

root.mainloop()